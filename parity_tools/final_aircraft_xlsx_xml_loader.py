#!/usr/bin/env python3
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text == "":
        return None
    try:
        numeric = float(text)
        return int(numeric) if numeric.is_integer() else numeric
    except ValueError:
        return text


def extract_sheet_matrix(ws):
    rows = ws.max_row or 1
    cols = ws.max_column or 1
    matrix = [[None for _ in range(cols)] for _ in range(rows)]
    for r_idx, row in enumerate(ws.iter_rows(), start=0):
        for c_idx, cell in enumerate(row, start=0):
            matrix[r_idx][c_idx] = normalize_value(cell.value)
    return matrix


def main():
    if len(sys.argv) != 2:
        print("Usage: final_aircraft_xlsx_xml_loader.py <workbook>", file=sys.stderr)
        sys.exit(2)

    workbook_path = Path(sys.argv[1])
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    wanted = {
        "aero": "Aero",
        "miss": "Miss",
        "main": "Main",
        "consts": "Consts",
        "gear": "Gear",
        "geom": "Geom",
    }

    sheets = {}
    for out_name, sheet_name in wanted.items():
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f'Required sheet "{sheet_name}" is missing.')
        sheets[out_name] = extract_sheet_matrix(wb[sheet_name])

    payload = {
        "fileName": workbook_path.name,
        "sheets": sheets,
    }
    print(json.dumps(payload))


if __name__ == "__main__":
    main()
